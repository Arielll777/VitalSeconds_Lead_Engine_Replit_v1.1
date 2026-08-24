"""v1.1 audit items: campaign vs technical, unknown export leak, release/resume, Grisha, unique active, append-only, backup, sqlite gate, historical master."""

from __future__ import annotations

import sqlite3

import pytest
from openpyxl import Workbook

from vitalseconds.db.migrations import run_migrations
from vitalseconds.db.session import DatabaseConfigError, DbConnection, get_connection
from vitalseconds.services.batch import BatchService
from vitalseconds.services.candidate_generator import CandidateGenerator
from vitalseconds.services.exporter import BackupError, ExportService
from vitalseconds.services.importer import MasterImporter, interpret_historical_status
from vitalseconds.services.master_workbook import inspect_workbook
from vitalseconds.services.override import OverrideService
from vitalseconds.services.restore import RestoreError, commit_restore, preview_backup, validate_backup
from vitalseconds.services.waterfall import WaterfallEngine
from tests.helpers import open_sqlite, seed_company_domain_exec, seed_exec_with_candidates


def test_already_contacted_does_not_rewrite_technical(db):
    BatchService(db).create("Batch_V", "NEVERBOUNCE")
    eid, _ = seed_exec_with_candidates(db)
    engine = WaterfallEngine(db)
    active = db.execute("SELECT email FROM email_candidates WHERE is_active = 1").fetchone()["email"]
    engine.process_result(active, "valid", "Batch_V", "nb.csv")
    res = OverrideService(db).mark_already_contacted(eid, "spoke last week")
    assert res["ok"] is True
    row = db.execute(
        "SELECT technical_status, campaign_disposition FROM executives WHERE executive_id = ?", (eid,)
    ).fetchone()
    assert row["technical_status"] == "VERIFIED_VALID"
    assert row["campaign_disposition"] == "ALREADY_CONTACTED"


def test_valid_plus_dnc_excluded_from_grisha(db):
    BatchService(db).create("Batch_DNC", "NEVERBOUNCE")
    eid, _ = seed_exec_with_candidates(db)
    engine = WaterfallEngine(db)
    active = db.execute("SELECT email FROM email_candidates WHERE is_active = 1").fetchone()["email"]
    engine.process_result(active, "valid", "Batch_DNC", "nb.csv")
    OverrideService(db).mark_do_not_contact(eid, "board request")
    path, count = ExportService(db).preview_grisha_ready()
    assert count == 0
    row = db.execute(
        "SELECT technical_status, campaign_disposition FROM executives WHERE executive_id = ?", (eid,)
    ).fetchone()
    assert row["technical_status"] == "VERIFIED_VALID"
    assert row["campaign_disposition"] == "DO_NOT_CONTACT"


def test_unknown_not_in_normal_pass_export(db):
    BatchService(db).create("Batch_U", "NEVERBOUNCE")
    eid, _ = seed_exec_with_candidates(db)
    engine = WaterfallEngine(db)
    active = db.execute("SELECT email FROM email_candidates WHERE is_active = 1").fetchone()["email"]
    engine.process_result(active, "unknown", "Batch_U", "nb.csv")
    path, count = ExportService(db).active_pass(include_unknown=False)
    assert count == 0
    path2, count2 = ExportService(db).unknown_retry()
    assert count2 == 1
    path3, count3 = ExportService(db).active_pass(include_unknown=True)
    assert count3 == 1


def test_release_holds_and_does_not_auto_resume(db):
    BatchService(db).create("Batch_AA", "NEVERBOUNCE")
    eid, did = seed_exec_with_candidates(db)
    engine = WaterfallEngine(db)
    active = db.execute("SELECT email FROM email_candidates WHERE is_active = 1").fetchone()["email"]
    engine.process_result(active, "accept_all", "Batch_AA", "nb.csv")
    domain = db.execute("SELECT normalized_domain FROM domains WHERE domain_id = ?", (did,)).fetchone()[0]
    res = OverrideService(db).release_quarantined_domain(domain, "false positive catch-all")
    assert res["ok"] is True
    dom = db.execute("SELECT technical_status FROM domains WHERE domain_id = ?", (did,)).fetchone()
    assert dom["technical_status"] == "RELEASED"
    ex = db.execute(
        "SELECT campaign_disposition, technical_status, active_candidate_id FROM executives WHERE executive_id = ?",
        (eid,),
    ).fetchone()
    assert ex["campaign_disposition"] == "HOLD"
    assert ex["active_candidate_id"] is None
    # still not in pass export
    _, count = ExportService(db).active_pass()
    assert count == 0


def test_resume_verification_explicit(db):
    BatchService(db).create("Batch_AA2", "NEVERBOUNCE")
    eid, did = seed_exec_with_candidates(db)
    engine = WaterfallEngine(db)
    active = db.execute("SELECT email FROM email_candidates WHERE is_active = 1").fetchone()["email"]
    engine.process_result(active, "accept_all", "Batch_AA2", "nb.csv")
    domain = db.execute("SELECT normalized_domain FROM domains WHERE domain_id = ?", (did,)).fetchone()[0]
    OverrideService(db).release_quarantined_domain(domain, "release")
    res = OverrideService(db).resume_verification(eid, "operator confirmed resume")
    assert res["ok"] is True
    ex = db.execute(
        "SELECT campaign_disposition, active_candidate_id FROM executives WHERE executive_id = ?",
        (eid,),
    ).fetchone()
    assert ex["campaign_disposition"] == "ACTIVE"
    assert ex["active_candidate_id"] is not None
    _, count = ExportService(db).active_pass()
    assert count == 1


def test_grisha_preview_does_not_mark_exported(db):
    BatchService(db).create("Batch_G", "NEVERBOUNCE")
    eid, _ = seed_exec_with_candidates(db)
    engine = WaterfallEngine(db)
    active = db.execute("SELECT email FROM email_candidates WHERE is_active = 1").fetchone()["email"]
    engine.process_result(active, "valid", "Batch_G", "nb.csv")
    _, n = ExportService(db).preview_grisha_ready()
    assert n == 1
    events = db.execute("SELECT COUNT(*) AS n FROM grisha_export_events").fetchone()[0]
    assert events == 0
    camp = db.execute("SELECT campaign_disposition FROM executives WHERE executive_id = ?", (eid,)).fetchone()[0]
    assert camp == "ACTIVE"


def test_grisha_confirm_records_event_with_batch_id(db):
    BatchService(db).create("Batch_G2", "NEVERBOUNCE")
    eid, _ = seed_exec_with_candidates(db)
    engine = WaterfallEngine(db)
    active = db.execute("SELECT email FROM email_candidates WHERE is_active = 1").fetchone()["email"]
    engine.process_result(active, "valid", "Batch_G2", "nb.csv")
    path, n = ExportService(db).confirm_grisha_export("Grisha_Export_TEST")
    assert n == 1
    ev = db.execute("SELECT batch_id, export_type, executive_id FROM grisha_export_events").fetchone()
    assert ev["batch_id"] == "Grisha_Export_TEST"
    assert ev["export_type"] == "GRISHA_READY"
    assert ev["executive_id"] == eid
    camp = db.execute("SELECT campaign_disposition FROM executives WHERE executive_id = ?", (eid,)).fetchone()[0]
    assert camp == "EXPORTED_TO_GRISHA"
    _, n2 = ExportService(db).preview_grisha_ready()
    assert n2 == 0


def test_one_active_candidate_database_enforced(db):
    eid, did = seed_exec_with_candidates(db)
    # two pending exist; forcing a second active must fail unique index
    rows = db.execute(
        "SELECT candidate_id FROM email_candidates WHERE executive_id = ? ORDER BY attempt_order",
        (eid,),
    ).fetchall()
    assert len(rows) >= 2
    second = rows[1]["candidate_id"]
    with pytest.raises(Exception):
        db.execute(
            "UPDATE email_candidates SET is_active = 1 WHERE candidate_id = ?",
            (second,),
        )
        db.commit()


def test_append_only_verification_events(db):
    BatchService(db).create("Batch_AO", "NEVERBOUNCE")
    seed_exec_with_candidates(db)
    engine = WaterfallEngine(db)
    active = db.execute("SELECT email FROM email_candidates WHERE is_active = 1").fetchone()["email"]
    engine.process_result(active, "unknown", "Batch_AO", "nb.csv")
    db.commit()
    eid_row = db.execute("SELECT event_id FROM verification_events").fetchone()
    assert eid_row is not None
    with pytest.raises(Exception):
        db.execute(
            "UPDATE verification_events SET neverbounce_status = 'valid' WHERE event_id = ?",
            (eid_row["event_id"],),
        )
    db.rollback()
    still = db.execute(
        "SELECT neverbounce_status FROM verification_events WHERE event_id = ?",
        (eid_row["event_id"],),
    ).fetchone()
    assert still["neverbounce_status"] == "unknown"
    with pytest.raises(Exception):
        db.execute("DELETE FROM verification_events WHERE event_id = ?", (eid_row["event_id"],))
    db.rollback()
    n = db.execute("SELECT COUNT(*) AS n FROM verification_events").fetchone()[0]
    assert n == 1


def test_backup_fails_if_required_table_missing(db, tmp_path):
    class Boom(DbConnection):
        def __init__(self, inner):
            self.backend = inner.backend
            self._raw = inner._raw
            self.inner = inner

        def execute(self, sql, params=None):
            if "SELECT * FROM audit_log" in sql:
                raise RuntimeError("relation audit_log does not exist")
            return self.inner.execute(sql, params)

    exporter = ExportService(Boom(db), export_dir=tmp_path)
    with pytest.raises(BackupError) as ei:
        exporter.full_state_backup()
    assert "audit_log" in str(ei.value)


def test_full_backup_roundtrip_restore_requires_confirm(db, tmp_path):
    seed_company_domain_exec(db)
    exporter = ExportService(db, export_dir=tmp_path)
    path, total = exporter.full_state_backup()
    assert path.exists()
    content = path.read_bytes()
    preview = preview_backup(content)
    assert preview["valid"] is True
    check = validate_backup(content)
    assert check["ok"] is True
    with pytest.raises(RestoreError):
        commit_restore(db, content, confirm_overwrite=False, batch_id="Restore_1")


def test_sqlite_forbidden_without_flag(monkeypatch):
    import vitalseconds.db.session as session_mod

    monkeypatch.setattr(session_mod, "DATABASE_URL", "")
    monkeypatch.setattr(session_mod, "ALLOW_SQLITE_TEST", False)
    with pytest.raises(DatabaseConfigError, match="DATABASE_URL is required"):
        get_connection()


def test_historical_valid_not_rewritten_as_pending_public_exact(db):
    importer = MasterImporter(db)
    stats = importer.import_rows(
        [
            {
                "first_name": "John",
                "last_name": "Smith",
                "company": "Life Star EMS",
                "domain": "lifestarems.com",
                "email": "jsmith@lifestarems.com",
                "public_email": "jsmith@lifestarems.com",
                "master_status": "VALID",
            }
        ],
        "Batch_HIST",
    )
    db.commit()
    row = db.execute(
        "SELECT technical_status, campaign_disposition FROM executives"
    ).fetchone()
    assert row["technical_status"] == "VERIFIED_VALID"
    cand = db.execute("SELECT status, candidate_basis, attempt_order FROM email_candidates").fetchone()
    assert cand["status"] == "VERIFIED_VALID"
    assert cand["status"] != "PENDING"
    assert cand["candidate_basis"] != "PUBLIC_EXACT" or cand["status"] == "VERIFIED_VALID"


def test_interpret_unknown_and_accept_all():
    u = interpret_historical_status({"master_status": "UNKNOWN"})
    assert u["technical_status"] == "UNKNOWN_PENDING_RETRY"
    a = interpret_historical_status({"master_status": "ACCEPT_ALL"})
    assert a["technical_status"] == "DOMAIN_QUARANTINED"


def test_master_workbook_detects_all_sheets_not_just_first(tmp_path):
    wb = Workbook()
    # first sheet is NOT the only one that matters
    wb.active.title = "1_Sparke_Raw_Input"
    wb.active.append(["first_name", "last_name", "company", "domain"])
    wb.active.append(["A", "B", "Co", "co.com"])
    s2 = wb.create_sheet("2_Verified_Ready_For_Grisha")
    s2.append(["first_name", "last_name", "company", "domain", "email"])
    s2.append(["C", "D", "Co2", "co2.com", "c@co2.com"])
    wb.create_sheet("mystery_tab")
    wb.create_sheet("8_Gemini_Instructions")
    path = tmp_path / "master.xlsx"
    wb.save(path)
    inspection = inspect_workbook(path.read_bytes(), "master.xlsx")
    names = [s["sheet_name"] for s in inspection["sheets"]]
    assert names[0] == "1_Sparke_Raw_Input"
    assert "2_Verified_Ready_For_Grisha" in names
    assert "mystery_tab" in inspection["unrecognized"]
    instr = next(s for s in inspection["sheets"] if s["sheet_name"] == "8_Gemini_Instructions")
    assert instr["skip"] is True
    assert instr["will_import"] is False
    mystery = next(s for s in inspection["sheets"] if s["sheet_name"] == "mystery_tab")
    assert mystery["will_import"] is False
    assert inspection["sheet_count"] == 4


def test_persistence_reopen_sqlite(db_path):
    conn = open_sqlite(db_path)
    seed_company_domain_exec(conn)
    conn.commit()
    n = conn.execute("SELECT COUNT(*) AS n FROM companies").fetchone()[0]
    conn.close()
    conn2 = open_sqlite(db_path)
    n2 = conn2.execute("SELECT COUNT(*) AS n FROM companies").fetchone()[0]
    conn2.close()
    assert n == 1
    assert n2 == 1


def test_migration_duplicate_active_goes_to_review_not_deleted(tmp_path):
    """Simulate dirty pre-index data: two actives, then run migrations 3+5."""
    path = tmp_path / "dirty.db"
    raw = sqlite3.connect(str(path))
    raw.row_factory = sqlite3.Row
    raw.executescript(
        """
        CREATE TABLE companies (
            company_id INTEGER PRIMARY KEY, canonical_name TEXT, normalized_name TEXT,
            buying_group TEXT, current_status TEXT DEFAULT 'ACTIVE',
            merged_into_company_id INTEGER, notes TEXT,
            created_at TEXT DEFAULT (datetime('now')), updated_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE executives (
            executive_id INTEGER PRIMARY KEY, company_id INTEGER, first_name TEXT, last_name TEXT,
            normalized_full_name TEXT, title TEXT, state TEXT,
            technical_status TEXT DEFAULT 'NET_NEW', campaign_disposition TEXT DEFAULT 'ACTIVE',
            current_status TEXT DEFAULT 'NET_NEW', primary_classification TEXT,
            active_candidate_id INTEGER,
            created_at TEXT DEFAULT (datetime('now')), updated_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE email_candidates (
            candidate_id INTEGER PRIMARY KEY, executive_id INTEGER, email TEXT,
            normalized_email TEXT UNIQUE, domain_id INTEGER, candidate_basis TEXT,
            pass_order INTEGER DEFAULT 0, attempt_order INTEGER DEFAULT 0,
            is_active INTEGER DEFAULT 0, status TEXT DEFAULT 'PENDING',
            created_at TEXT DEFAULT (datetime('now'))
        );
        CREATE TABLE audit_log (
            audit_id INTEGER PRIMARY KEY, entity_type TEXT, entity_id TEXT, action TEXT,
            old_value TEXT, new_value TEXT, reason TEXT, batch_id TEXT, user_context TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );
        INSERT INTO companies (company_id, canonical_name, normalized_name) VALUES (1, 'X', 'x');
        INSERT INTO executives (executive_id, company_id, first_name, last_name, normalized_full_name)
        VALUES (1, 1, 'A', 'B', 'a b');
        INSERT INTO email_candidates (candidate_id, executive_id, email, normalized_email, candidate_basis, is_active)
        VALUES (1, 1, 'a@x.com', 'a@x.com', 'P', 1);
        INSERT INTO email_candidates (candidate_id, executive_id, email, normalized_email, candidate_basis, is_active)
        VALUES (2, 1, 'b@x.com', 'b@x.com', 'P', 1);
        """
    )
    raw.commit()
    conn = DbConnection("sqlite", raw)
    # Pretend 001 already applied so we don't recreate; run 003+ on existing tables via full run
    # schema_version empty → 001 CREATE IF NOT EXISTS then 003 reviews duplicates
    run_migrations(conn)
    conn.commit()
    actives = conn.execute(
        "SELECT candidate_id, is_active, status FROM email_candidates ORDER BY candidate_id"
    ).fetchall()
    statuses = {r["candidate_id"]: (r["is_active"], r["status"]) for r in actives}
    assert statuses[1][0] == 1
    assert statuses[2][0] == 0
    assert statuses[2][1] == "IMPORT_REVIEW"
    still = conn.execute("SELECT COUNT(*) AS n FROM email_candidates").fetchone()[0]
    assert still == 2  # never deleted
    camp = conn.execute("SELECT campaign_disposition FROM executives WHERE executive_id = 1").fetchone()[0]
    assert camp == "HOLD"
    conn.close()
