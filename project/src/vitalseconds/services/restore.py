"""Restore Full State Backup XLSX. PREVIEW → VALIDATE → COMMIT. Never silent overwrite."""

from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List

from openpyxl import load_workbook

from vitalseconds.config import REQUIRED_BACKUP_TABLES


class RestoreError(RuntimeError):
    pass


def preview_backup(content: bytes) -> Dict[str, Any]:
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        sheets[name] = {
            "row_count": max(0, len(rows) - 1) if rows else 0,
            "headers": [str(h) for h in rows[0]] if rows else [],
        }
    missing = [t for t in REQUIRED_BACKUP_TABLES if t not in sheets]
    return {
        "sheets": sheets,
        "missing_required": missing,
        "valid": len(missing) == 0,
    }


def validate_backup(content: bytes) -> Dict[str, Any]:
    preview = preview_backup(content)
    errors: List[str] = []
    if preview["missing_required"]:
        errors.append("Missing required sheets: " + ", ".join(preview["missing_required"]))
    return {"ok": not errors, "errors": errors, "preview": preview}


def commit_restore(conn, content: bytes, *, confirm_overwrite: bool, batch_id: str) -> Dict[str, Any]:
    if not confirm_overwrite:
        raise RestoreError("Explicit confirm_overwrite=True is required to restore into an operational database.")
    check = validate_backup(content)
    if not check["ok"]:
        raise RestoreError("Backup validation failed: " + "; ".join(check["errors"]))

    existing = conn.execute("SELECT COUNT(*) AS c FROM companies").fetchone()[0]
    if existing and not confirm_overwrite:
        raise RestoreError("Operational database is not empty.")

    wb = load_workbook(BytesIO(content), data_only=True)
    restored = {}
    # Insert in FK-safe order (REQUIRED_BACKUP_TABLES is already ordered)
    for table in REQUIRED_BACKUP_TABLES:
        ws = wb[table]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            restored[table] = 0
            continue
        headers = [str(h) for h in rows[0]]
        if headers == ["(empty)"]:
            restored[table] = 0
            continue
        count = 0
        for data in rows[1:]:
            if all(v is None or v == "" for v in data):
                continue
            placeholders = ", ".join("?" for _ in headers)
            cols = ", ".join(headers)
            values = ["" if v is None else v for v in data[: len(headers)]]
            try:
                conn.execute(
                    f"INSERT OR IGNORE INTO {table} ({cols}) VALUES ({placeholders})",
                    tuple(values),
                )
                count += 1
            except Exception as exc:
                raise RestoreError(f"Restore failed on {table}: {exc}") from exc
        restored[table] = count
    conn.execute(
        """
        INSERT INTO audit_log (entity_type, entity_id, action, new_value, reason, batch_id)
        VALUES ('RESTORE', ?, 'FULL_STATE_RESTORE', ?, 'Confirmed restore from XLSX backup', ?)
        """,
        (batch_id, str(restored), batch_id),
    )
    return {"restored": restored, "batch_id": batch_id}
